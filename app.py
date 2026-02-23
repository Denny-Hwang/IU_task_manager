import streamlit as st
from streamlit_echarts import st_echarts
from datetime import date, timedelta
import json
import os
import io
import html as html_mod
import unicodedata

# ─── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Quadrant Task Manager",
    page_icon="📊",
    layout="wide",
)

# ─── Constants ─────────────────────────────────────────────────────────────────
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tasks_data.json")

QUADRANT_META = {
    "do_first": {
        "bg": "rgba(239,68,68,0.10)",
        "color": "#DC2626",
        "color_alpha": "rgba(220,38,38,0.3)",
        "fill": "FFCCCC",
        "label": "Do First",
        "sub": "Urgent & Important",
        "emoji": "🔴",
    },
    "delegate": {
        "bg": "rgba(249,115,22,0.10)",
        "color": "#EA580C",
        "color_alpha": "rgba(234,88,12,0.3)",
        "fill": "FFE0CC",
        "label": "Delegate",
        "sub": "Urgent & Less Important",
        "emoji": "🟠",
    },
    "schedule": {
        "bg": "rgba(59,130,246,0.10)",
        "color": "#2563EB",
        "color_alpha": "rgba(37,99,235,0.3)",
        "fill": "CCE0FF",
        "label": "Schedule",
        "sub": "Less Urgent & Important",
        "emoji": "🔵",
    },
    "eliminate": {
        "bg": "rgba(34,197,94,0.10)",
        "color": "#16A34A",
        "color_alpha": "rgba(22,163,74,0.3)",
        "fill": "CCFFCC",
        "label": "Eliminate",
        "sub": "Less Urgent & Less Important",
        "emoji": "🟢",
    },
}


# ─── Data Helpers ──────────────────────────────────────────────────────────────
def load_tasks():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []


def save_tasks(tasks):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(tasks, f, ensure_ascii=False, indent=2)


def calc_urgency(deadline_str: str) -> float:
    """Map deadline to urgency 1-10. Closer deadline = higher urgency."""
    days = (date.fromisoformat(deadline_str) - date.today()).days
    if days <= 0:
        return 10.0
    if days >= 30:
        return 1.0
    return round(10.0 - days * 9.0 / 30.0, 1)


def quadrant_key(importance: float, urgency: float) -> str:
    if importance > 5 and urgency > 5:
        return "do_first"
    if importance <= 5 and urgency > 5:
        return "delegate"
    if importance > 5 and urgency <= 5:
        return "schedule"
    return "eliminate"


def quadrant_info(importance, urgency):
    return QUADRANT_META[quadrant_key(importance, urgency)]


def _d_day_text(deadline_str):
    days = (date.fromisoformat(deadline_str) - date.today()).days
    return f"D-{days}" if days >= 0 else f"D+{abs(days)}"


# ─── Text Helpers ──────────────────────────────────────────────────────────────
def _display_width(s):
    w = 0
    for ch in s:
        if unicodedata.east_asian_width(ch) in ("W", "F"):
            w += 2
        else:
            w += 1
    return w


def _truncate(s, max_w):
    w = 0
    for i, ch in enumerate(s):
        cw = 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
        if w + cw > max_w:
            return s[:i] + "…"
        w += cw
    return s


def format_block_label(name, deadline_str):
    """Format task name + D-day to fit inside chart block (2 lines, newline separated)."""
    MAX_WIDTH = 12
    d_text = _d_day_text(deadline_str)

    if _display_width(name) <= MAX_WIDTH:
        return f"{name}\n{d_text}"

    candidates = []
    for i, ch in enumerate(name):
        if ch == " " and 0 < i < len(name) - 1:
            lw = _display_width(name[:i])
            rw = _display_width(name[i + 1:])
            if lw <= MAX_WIDTH:
                candidates.append((i, abs(lw - rw)))

    if candidates:
        best_break = min(candidates, key=lambda x: (x[1], -x[0]))[0]
        line1 = name[:best_break]
        line2 = name[best_break + 1:]
        if _display_width(line2) > MAX_WIDTH:
            line2 = _truncate(line2, MAX_WIDTH - 1)
        return f"{line1}\n{line2}"

    return f"{_truncate(name, MAX_WIDTH)}\n{d_text}"


def _task_summary_text(task):
    """Build plain-text summary for clipboard copy."""
    qi = quadrant_info(task["importance"], task["urgency"])
    d_text = _d_day_text(task["deadline"])
    lines = [
        f"Task: {task['name']}",
        f"Importance: {task['importance']}/10",
        f"Urgency: {task['urgency']:.1f}/10",
        f"Deadline: {task['deadline']} ({d_text})",
        f"Quadrant: {qi['emoji']} {qi['label']}",
    ]
    if task.get("description"):
        lines.append(f"Description: {task['description']}")
    return "\n".join(lines)


# ─── Example Tasks ─────────────────────────────────────────────────────────────
def generate_example_tasks():
    today = date.today()
    examples = [
        {
            "name": "Server Outage Response",
            "description": "Production server emergency — CPU usage exceeded 95%. Immediate diagnosis and resolution required.",
            "importance": 9,
            "deadline": (today + timedelta(days=1)).isoformat(),
            "color": "#E74C3C",
        },
        {
            "name": "Client Meeting Prep",
            "description": "Prepare Thursday client presentation. Finalize proposal and demo scenario.",
            "importance": 8,
            "deadline": (today + timedelta(days=3)).isoformat(),
            "color": "#C0392B",
        },
        {
            "name": "Reply to Emails",
            "description": "Reply to manager's email and cross-team inquiries. Simple confirm/forward tasks.",
            "importance": 3,
            "deadline": (today + timedelta(days=2)).isoformat(),
            "color": "#F39C12",
        },
        {
            "name": "Submit Expense Report",
            "description": "Submit last month's travel and expense report to the finance team.",
            "importance": 4,
            "deadline": (today + timedelta(days=4)).isoformat(),
            "color": "#E67E22",
        },
        {
            "name": "New Feature Design",
            "description": "Q2 roadmap core feature architecture design. Includes tech stack selection.",
            "importance": 9,
            "deadline": (today + timedelta(days=21)).isoformat(),
            "color": "#3498DB",
        },
        {
            "name": "Tech Debt Cleanup",
            "description": "Plan legacy code refactoring and test coverage improvement.",
            "importance": 7,
            "deadline": (today + timedelta(days=18)).isoformat(),
            "color": "#2980B9",
        },
        {
            "name": "Organize Supplies",
            "description": "Tidy desk and discard unnecessary documents. Handle when free.",
            "importance": 2,
            "deadline": (today + timedelta(days=25)).isoformat(),
            "color": "#2ECC71",
        },
        {
            "name": "Unsubscribe Newsletters",
            "description": "Cancel unread newsletter subscriptions and clean up info channels.",
            "importance": 1,
            "deadline": (today + timedelta(days=28)).isoformat(),
            "color": "#1ABC9C",
        },
    ]
    for t in examples:
        t["urgency"] = calc_urgency(t["deadline"])
    return examples


# ─── Export: Excel with Gantt & Milestones ─────────────────────────────────────
def build_excel(tasks):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Sheet 1: Task Summary
    ws1 = wb.active
    ws1.title = "Tasks"
    headers1 = ["Task Name", "Description", "Importance", "Urgency", "Deadline", "D-Day", "Quadrant"]
    for j, h in enumerate(headers1, 1):
        cell = ws1.cell(1, j, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, task in enumerate(tasks, 2):
        qi = quadrant_info(task["importance"], task["urgency"])
        d_text = _d_day_text(task["deadline"])
        row = [
            task["name"],
            task.get("description", ""),
            task["importance"],
            round(task["urgency"], 1),
            task["deadline"],
            d_text,
            f"{qi['emoji']} {qi['label']}",
        ]
        for j, val in enumerate(row, 1):
            cell = ws1.cell(i, j, val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(j == 2))

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 16

    # ── Sheet 2: Gantt Chart
    ws2 = wb.create_sheet("Gantt Chart")
    today = date.today()
    if tasks:
        deadlines = [date.fromisoformat(t["deadline"]) for t in tasks]
        end_date = max(max(deadlines), today) + timedelta(days=3)
    else:
        end_date = today + timedelta(days=30)

    date_range = []
    cur = today
    while cur <= end_date:
        date_range.append(cur)
        cur += timedelta(days=1)

    for j, h in enumerate(["Task", "Quadrant", "Deadline"], 1):
        cell = ws2.cell(1, j, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    from openpyxl.utils import get_column_letter

    today_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for j, d in enumerate(date_range):
        col = j + 4
        cell = ws2.cell(1, col, d.strftime("%m/%d"))
        cell.font = Font(bold=(d == today), size=9, color="FFFFFF" if d != today else "000000")
        cell.fill = header_fill if d != today else today_fill
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col)].width = 5

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12

    for i, task in enumerate(sorted(tasks, key=lambda t: t["deadline"]), 2):
        qk = quadrant_key(task["importance"], task["urgency"])
        qi = QUADRANT_META[qk]
        fill = PatternFill(start_color=qi["fill"], end_color=qi["fill"], fill_type="solid")
        dl = date.fromisoformat(task["deadline"])
        ws2.cell(i, 1, task["name"]).border = thin_border
        ws2.cell(i, 2, qi["label"]).border = thin_border
        ws2.cell(i, 3, task["deadline"]).border = thin_border
        for j, d in enumerate(date_range):
            col = j + 4
            cell = ws2.cell(i, col)
            cell.border = thin_border
            if d <= dl:
                cell.fill = fill
            if d == dl:
                cell.value = "▶"
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal="center")
            if d == today:
                cell.border = Border(
                    left=Side(style="medium", color="FF8C00"),
                    right=Side(style="medium", color="FF8C00"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )

    # ── Sheet 3: Milestones
    ws3 = wb.create_sheet("Milestones")
    for j, h in enumerate(["#", "Milestone", "Deadline", "D-Day", "Status", "Importance", "Urgency", "Quadrant"], 1):
        cell = ws3.cell(1, j, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    status_fills = {
        "overdue": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "today": PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid"),
        "week": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "upcoming": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    }
    for i, task in enumerate(sorted(tasks, key=lambda t: t["deadline"]), 2):
        qi = quadrant_info(task["importance"], task["urgency"])
        days_left = (date.fromisoformat(task["deadline"]) - today).days
        d_text = _d_day_text(task["deadline"])
        if days_left < 0:
            status, sfill = "Overdue", status_fills["overdue"]
        elif days_left == 0:
            status, sfill = "Due Today", status_fills["today"]
        elif days_left <= 7:
            status, sfill = "This Week", status_fills["week"]
        else:
            status, sfill = "Upcoming", status_fills["upcoming"]
        row = [i - 1, task["name"], task["deadline"], d_text, status,
               task["importance"], round(task["urgency"], 1), f"{qi['emoji']} {qi['label']}"]
        for j, val in enumerate(row, 1):
            cell = ws3.cell(i, j, val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if j == 5:
                cell.fill = sfill

    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 12
    ws3.column_dimensions["G"].width = 10
    ws3.column_dimensions["H"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_chart_html(chart_option):
    """Generate standalone ECharts HTML page from chart option dict."""
    option_json = json.dumps(chart_option, ensure_ascii=False).replace("</", "<\\/")
    return (
        '<!DOCTYPE html>\n'
        '<html lang="en">\n'
        '<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        '<title>Quadrant Task Manager</title>\n'
        '<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>\n'
        '<style>body{margin:0;padding:0;background:#fff}#chart{width:100vw;height:100vh}</style>\n'
        '</head>\n'
        '<body>\n'
        '<div id="chart"></div>\n'
        '<script>\n'
        'var chart=echarts.init(document.getElementById("chart"));\n'
        f'chart.setOption({option_json});\n'
        'window.addEventListener("resize",function(){chart.resize()});\n'
        '</script>\n'
        '</body>\n'
        '</html>'
    )


# ─── Session State ─────────────────────────────────────────────────────────────
if "initialized" not in st.session_state:
    st.session_state.initialized = True
    existing = load_tasks()
    if existing:
        st.session_state.tasks = existing
        st.session_state.demo_mode = False
    else:
        st.session_state.tasks = generate_example_tasks()
        st.session_state.demo_mode = True

for _t in st.session_state.tasks:
    _t["urgency"] = calc_urgency(_t["deadline"])


# ─── Helper: show task detail panel ───────────────────────────────────────────
def _show_task_detail(task, idx, tasks_list):
    """Render task detail panel with metrics, edit controls, and clipboard copy."""
    qi = quadrant_info(task["importance"], task["urgency"])
    days_left = (date.fromisoformat(task["deadline"]) - date.today()).days

    st.divider()
    st.subheader(f"{qi['emoji']} {task['name']}")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Importance", f"{task['importance']} / 10")
    with c2:
        st.metric("Urgency", f"{task['urgency']:.1f} / 10")
    with c3:
        if days_left >= 0:
            st.metric("D-Day", task["deadline"], delta=f"D-{days_left}")
        else:
            st.metric("D-Day", task["deadline"],
                      delta=f"{abs(days_left)} days overdue", delta_color="inverse")
    with c4:
        st.metric("Quadrant", qi["label"])

    if task.get("description"):
        st.info(task["description"])
    else:
        st.caption("No description.")

    # ── Edit position (Importance + Deadline → Urgency auto-recalc) ────────
    st.markdown("**✏️ Edit Position** — adjust values below, then Update or Revert")
    ea, eb = st.columns(2)
    with ea:
        new_imp = st.slider("Importance", 1, 10, task["importance"], key=f"eimp_{idx}")
    with eb:
        new_deadline = st.date_input("Deadline", value=date.fromisoformat(task["deadline"]), key=f"edl_{idx}")

    new_urg = calc_urgency(new_deadline.isoformat())
    imp_changed = new_imp != task["importance"]
    dl_changed = new_deadline.isoformat() != task["deadline"]

    if imp_changed or dl_changed:
        new_qi = quadrant_info(new_imp, new_urg)
        st.warning(
            f"Preview: Importance {new_imp} · Urgency {new_urg:.1f} "
            f"→ {new_qi['emoji']} **{new_qi['label']}**"
        )
        uc, rc = st.columns(2)
        with uc:
            if st.button("✅ Update", type="primary", use_container_width=True, key="btn_update"):
                tasks_list[idx]["importance"] = new_imp
                tasks_list[idx]["deadline"] = new_deadline.isoformat()
                tasks_list[idx]["urgency"] = new_urg
                save_tasks(tasks_list)
                for k in [f"eimp_{idx}", f"edl_{idx}"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with rc:
            if st.button("↩️ Revert", use_container_width=True, key="btn_revert"):
                for k in [f"eimp_{idx}", f"edl_{idx}"]:
                    st.session_state.pop(k, None)
                st.rerun()

    # ── Clipboard copy ─────────────────────────────────────────────────────
    st.markdown("**📋 Copy to clipboard** — click the copy icon (top-right of the box)")
    st.code(_task_summary_text(task), language=None)


# ─── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("➕ New Task")

    with st.form("add_task_form", clear_on_submit=True):
        name = st.text_input("Task Name")
        desc = st.text_area("Description", height=80)
        importance = st.slider("Importance", 1, 10, 5)
        deadline = st.date_input("Deadline", value=date.today() + timedelta(days=7))
        color = st.color_picker("Block Color", "#4A90D9")
        submitted = st.form_submit_button("Add Task", use_container_width=True)

        if submitted:
            if name.strip():
                if st.session_state.demo_mode:
                    st.session_state.demo_mode = False
                st.session_state.tasks.append({
                    "name": name.strip(),
                    "description": desc,
                    "importance": importance,
                    "deadline": deadline.isoformat(),
                    "urgency": calc_urgency(deadline.isoformat()),
                    "color": color,
                })
                save_tasks(st.session_state.tasks)
                st.rerun()
            else:
                st.error("Task name is required.")

    # ── Task list (expandable with details)
    st.divider()
    st.subheader(f"📋 Tasks ({len(st.session_state.tasks)})")

    for i, task in enumerate(st.session_state.tasks):
        qi = quadrant_info(task["importance"], task["urgency"])
        d_text = _d_day_text(task["deadline"])

        with st.expander(f"{qi['emoji']} {task['name']}  ·  {d_text}"):
            st.caption(
                f"Imp {task['importance']} · Urg {task['urgency']:.1f} · "
                f"Deadline: {task['deadline']}"
            )
            if task.get("description"):
                st.markdown(f"*{task['description']}*")
            else:
                st.caption("No description.")

            sc1, sc2 = st.columns(2)
            with sc1:
                if st.button("📌 Select", key=f"sel_{i}", use_container_width=True):
                    st.session_state.selected_task_idx = i
                    st.rerun()
            with sc2:
                if st.button("🗑 Delete", key=f"del_{i}", use_container_width=True):
                    st.session_state.tasks.pop(i)
                    if "selected_task_idx" in st.session_state:
                        del st.session_state["selected_task_idx"]
                    save_tasks(st.session_state.tasks)
                    st.rerun()

    # ── Reset Section
    st.divider()
    rc1, rc2 = st.columns(2)
    with rc1:
        if st.button("🗑️ Clear All", use_container_width=True, help="Delete all tasks"):
            st.session_state.tasks = []
            st.session_state.demo_mode = False
            if "selected_task_idx" in st.session_state:
                del st.session_state["selected_task_idx"]
            save_tasks([])
            st.rerun()
    with rc2:
        if st.button("📦 Load Examples", use_container_width=True, help="Load example tasks"):
            st.session_state.tasks = generate_example_tasks()
            st.session_state.demo_mode = True
            if "selected_task_idx" in st.session_state:
                del st.session_state["selected_task_idx"]
            if os.path.exists(DATA_FILE):
                os.remove(DATA_FILE)
            st.rerun()


# ─── Main Area ─────────────────────────────────────────────────────────────────
st.title("📊 Quadrant Task Manager")
st.caption("Importance × Urgency — Tap or click a task block to view details")

# ── Build ECharts (CHART FIRST) ───────────────────────────────────────────────
tasks = st.session_state.tasks

scatter_data = []
for t in tasks:
    qi = quadrant_info(t["importance"], t["urgency"])
    label_text = format_block_label(t["name"], t["deadline"])
    d_text = _d_day_text(t["deadline"])
    safe_name = html_mod.escape(t["name"])

    # Tooltip text — use \n (rendered via white-space: pre-line CSS)
    tooltip_text = (
        f"{safe_name}\n"
        f"Importance: {t['importance']}/10 · Urgency: {t['urgency']:.1f}/10\n"
        f"Deadline: {t['deadline']} ({d_text})\n"
        f"{qi['emoji']} {qi['label']}"
    )

    scatter_data.append({
        "value": [t["importance"], t["urgency"]],
        "name": tooltip_text,
        "itemStyle": {
            "color": t.get("color", "#4A90D9"),
            "borderColor": "#fff",
            "borderWidth": 2,
            "opacity": 0.9,
        },
        "label": {
            "show": True,
            "position": "inside",
            "formatter": label_text,
            "color": "#fff",
            "fontSize": 10,
            "fontWeight": "bold",
            "lineHeight": 14,
            "width": 72,
            "overflow": "truncate",
        },
    })

# Quadrant background labels
q_label_data = []
_positions = {
    "do_first": [7.75, 7.75],
    "delegate": [2.5, 7.75],
    "schedule": [7.75, 2.5],
    "eliminate": [2.5, 2.5],
}
for qk, coord in _positions.items():
    qm = QUADRANT_META[qk]
    q_label_data.append({
        "coord": coord,
        "symbol": "none",
        "label": {
            "show": True,
            "formatter": f"{qm['label']}\n{qm['sub']}",
            "color": qm["color_alpha"],
            "fontSize": 14,
            "fontWeight": "bold",
            "lineHeight": 20,
        },
    })

# Quadrant background areas
mark_area_data = [
    [{"xAxis": 5, "yAxis": 5, "itemStyle": {"color": QUADRANT_META["do_first"]["bg"], "borderWidth": 0}},
     {"xAxis": 10.5, "yAxis": 10.5}],
    [{"xAxis": 0, "yAxis": 5, "itemStyle": {"color": QUADRANT_META["delegate"]["bg"], "borderWidth": 0}},
     {"xAxis": 5, "yAxis": 10.5}],
    [{"xAxis": 5, "yAxis": 0, "itemStyle": {"color": QUADRANT_META["schedule"]["bg"], "borderWidth": 0}},
     {"xAxis": 10.5, "yAxis": 5}],
    [{"xAxis": 0, "yAxis": 0, "itemStyle": {"color": QUADRANT_META["eliminate"]["bg"], "borderWidth": 0}},
     {"xAxis": 5, "yAxis": 5}],
]

chart_option = {
    "backgroundColor": "#fff",
    "grid": {
        "left": 55,
        "right": 20,
        "top": 40,
        "bottom": 55,
        "containLabel": False,
    },
    "tooltip": {
        "trigger": "item",
        "formatter": "{b}",
        "backgroundColor": "rgba(255,255,255,0.96)",
        "borderColor": "#ddd",
        "textStyle": {"color": "#333", "fontSize": 13},
        "extraCssText": "box-shadow: 0 2px 8px rgba(0,0,0,0.12); max-width: 280px; white-space: pre-line;",
    },
    "xAxis": {
        "name": "Importance →",
        "nameLocation": "middle",
        "nameGap": 32,
        "nameTextStyle": {"fontSize": 14, "fontWeight": "bold", "color": "#555"},
        "min": 0,
        "max": 10.5,
        "interval": 1,
        "splitLine": {"lineStyle": {"color": "rgba(0,0,0,0.06)"}},
        "axisLine": {"lineStyle": {"color": "#888", "width": 2}},
        "axisTick": {"lineStyle": {"color": "#888", "width": 2}, "length": 6},
        "axisLabel": {"color": "#555", "fontSize": 13, "fontWeight": "bold"},
    },
    "yAxis": {
        "name": "Urgency →",
        "nameLocation": "middle",
        "nameGap": 40,
        "nameTextStyle": {"fontSize": 14, "fontWeight": "bold", "color": "#555"},
        "min": 0,
        "max": 10.5,
        "interval": 1,
        "splitLine": {"lineStyle": {"color": "rgba(0,0,0,0.06)"}},
        "axisLine": {"lineStyle": {"color": "#888", "width": 2}},
        "axisTick": {"lineStyle": {"color": "#888", "width": 2}, "length": 6},
        "axisLabel": {"color": "#555", "fontSize": 13, "fontWeight": "bold"},
    },
    "toolbox": {
        "right": 12,
        "top": 5,
        "feature": {
            "restore": {"show": True, "title": "Reset View"},
            "saveAsImage": {
                "show": True,
                "title": "Save as PNG",
                "pixelRatio": 2,
                "name": f"quadrant_chart_{date.today().isoformat()}",
            },
        },
        "iconStyle": {"borderColor": "#888"},
    },
    "dataZoom": [
        {"type": "inside", "xAxisIndex": 0},
        {"type": "inside", "yAxisIndex": 0},
    ],
    "series": [
        # Series 0: backgrounds, dividers, quadrant labels (non-interactive)
        {
            "type": "scatter",
            "data": [],
            "silent": True,
            "z": 1,
            "markArea": {"silent": True, "data": mark_area_data},
            "markLine": {
                "silent": True,
                "symbol": "none",
                "lineStyle": {"color": "rgba(0,0,0,0.15)", "width": 1.5, "type": "dashed"},
                "data": [{"xAxis": 5}, {"yAxis": 5}],
                "label": {"show": False},
            },
            "markPoint": {"silent": True, "data": q_label_data},
        },
        # Series 1: task blocks (interactive)
        {
            "type": "scatter",
            "symbol": "roundRect",
            "symbolSize": [80, 48],
            "data": scatter_data,
            "z": 2,
            "animationDuration": 500,
        },
    ],
}

# Render chart
events = {
    "click": "function(params) { return (params.seriesIndex === 1 && params.dataIndex !== undefined) ? params.dataIndex : -1; }"
}
clicked = st_echarts(chart_option, events=events, height="600px", key="quadrant_chart")

if clicked is not None:
    try:
        idx = int(clicked)
        if 0 <= idx < len(tasks):
            st.session_state.selected_task_idx = idx
    except (TypeError, ValueError):
        pass

# ── Clean up stale edit widget state on task switch ──────────────────────────
if "selected_task_idx" in st.session_state:
    _cur_idx = st.session_state.selected_task_idx
    _prev_idx = st.session_state.get("_prev_edit_idx", -1)
    if _cur_idx != _prev_idx:
        for _k in [f"eimp_{_prev_idx}", f"edl_{_prev_idx}"]:
            st.session_state.pop(_k, None)
        st.session_state._prev_edit_idx = _cur_idx

# ── Task Detail Panel (shown on chart click or sidebar click) ─────────────────
if tasks and "selected_task_idx" in st.session_state:
    idx = st.session_state.selected_task_idx
    if 0 <= idx < len(tasks):
        _show_task_detail(tasks[idx], idx, tasks)

elif not tasks:
    st.info("👈 Add your first task from the sidebar!")

# ── Export Section ─────────────────────────────────────────────────────────────
if tasks:
    st.divider()
    st.subheader("📥 Export")
    ex1, ex2 = st.columns(2)

    with ex1:
        excel_buf = build_excel(tasks)
        st.download_button(
            "📊 Excel (Gantt + Milestones)",
            data=excel_buf,
            file_name=f"task_plan_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with ex2:
        html_str = build_chart_html(chart_option)
        st.download_button(
            "🌐 Interactive Chart (HTML)",
            data=html_str,
            file_name=f"quadrant_chart_{date.today().isoformat()}.html",
            mime="text/html",
            use_container_width=True,
        )

    st.caption("💡 Use the toolbar icons (top-right of chart) to reset view or save as PNG.")

# ── Demo Mode Banner (below chart) ───────────────────────────────────────────
if st.session_state.demo_mode:
    st.divider()
    with st.container(border=True):
        st.markdown("#### 🎯 Demo Mode")
        st.markdown(
            "Example tasks are displayed on the chart. "
            "Click a block to see its details, or explore the quadrants below."
        )

        # 2x2 quadrant explanation grid
        st.markdown("##### Quadrant Guide")
        top1, top2 = st.columns(2)
        with top1:
            st.markdown(
                '<div style="background:rgba(249,115,22,0.08); border-left:4px solid #EA580C; '
                'padding:12px; border-radius:6px;">'
                '<b>🟠 Delegate</b><br/>'
                '<small>Urgent & Less Important</small><br/>'
                'Assign to others or handle quickly.</div>',
                unsafe_allow_html=True,
            )
        with top2:
            st.markdown(
                '<div style="background:rgba(239,68,68,0.08); border-left:4px solid #DC2626; '
                'padding:12px; border-radius:6px;">'
                '<b>🔴 Do First</b><br/>'
                '<small>Urgent & Important</small><br/>'
                'Act on these immediately.</div>',
                unsafe_allow_html=True,
            )
        bot1, bot2 = st.columns(2)
        with bot1:
            st.markdown(
                '<div style="background:rgba(34,197,94,0.08); border-left:4px solid #16A34A; '
                'padding:12px; border-radius:6px;">'
                '<b>🟢 Eliminate</b><br/>'
                '<small>Less Urgent & Less Important</small><br/>'
                'Consider dropping or postponing.</div>',
                unsafe_allow_html=True,
            )
        with bot2:
            st.markdown(
                '<div style="background:rgba(59,130,246,0.08); border-left:4px solid #2563EB; '
                'padding:12px; border-radius:6px;">'
                '<b>🔵 Schedule</b><br/>'
                '<small>Less Urgent & Important</small><br/>'
                'Plan ahead and block time.</div>',
                unsafe_allow_html=True,
            )

        st.markdown("")
        st.caption(
            "💡 **Urgency** is auto-calculated from the **Deadline**: "
            "D-0 → 10 (most urgent) / D-30+ → 1 (least urgent)"
        )

        bc1, bc2, _ = st.columns([1, 1, 3])
        with bc1:
            if st.button("🚀 **Start Fresh**", use_container_width=True, type="primary"):
                st.session_state.tasks = []
                st.session_state.demo_mode = False
                save_tasks([])
                st.rerun()
        with bc2:
            if st.button("✏️ Continue with Examples", use_container_width=True):
                st.session_state.demo_mode = False
                save_tasks(st.session_state.tasks)
                st.rerun()
